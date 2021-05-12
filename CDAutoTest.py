# -*- coding: utf-8 -*-

import xlrd
import json
import openpyxl
import numpy as np
import pandas as pd
import datetime
from datetime import timedelta
from fractions import Fraction


# 全局变量
start_time = ''
end_time = ''
spn_dict = {}
pgn_dict = {}
pgn_list = []


               
def HBFC_to_xlsx(type_cd):
    
    # 1、读取EXCEL并获取HB表
    testcasebook = openpyxl.load_workbook('CDTestCase.xlsx') # 读取xlsx文件
    table = testcasebook.get_sheet_by_name(type_cd) # 获得指定名称的页
    for cell in table['G']:
        cell.value = ''

    json_string = type_cd + '_CD.json'
    print(json_string)
    # 读取HB JSON
    with open(json_string, 'r', encoding='UTF-8') as f:
        data = json.load(f)
        for item in data:
            pname = item
            pvalue =data[item]
    #头信息第一次跟json对比
            for cell in table['B']:
                cell_name = str(cell.value).strip()
                cell_name = cell_name.strip('[]')

                if cell_name == pname:
            # 对比完把相应的值给写入
                    table.cell(row=cell.row, column=7).value = str(pvalue)
                    # table.cell(row=cell.row, column=7).value = str(pvalue)


            if item == 'Snapshots':
                snapDict = pvalue[0]
                for snapItem in snapDict:
                    snapItemValue = snapDict[snapItem]
                    pname = snapItem
                    pvalue = snapDict[snapItem]
                    if snapItem == 'Snapshot_DateTimestamp':
                        for cell in table['B']:
                            if 'Snapshot_DateTimestamp' in str(cell.value):
                                table.cell(row=cell.row, column=7).value = pvalue

            #快照参数第一次跟json对比
                    
                    if snapItem == 'Parameter':
                    #快照参数第二次跟json对比
                        for snapParameterDict in snapItemValue:
                        # print(snapParameterDict['Name'])
                        # print(snapParameterDict['Value'])
                            pname = snapParameterDict['Name']
                            pvalue = snapParameterDict['Value']
                            for cell in table['B']:
                        # print(cell.value)
                                if pname and cell.value:
                                    if pname.lower() in (cell.value.strip().lower()):
                                # 对比完把相应的值给写入
                                        table.cell(row=cell.row, column=7).value = pvalue

    # 保存excel
    testcasebook.save('CDTestCase.xlsx')


def PGNHexToDec(pgn, hexStr):
    sPGN = str(pgn)

    # print(len(hexStr))
    hexArray = np.array(hexStr.split())
    # print(len(hexArray))
    filename = 'rules.txt' # txt文件和当前脚本在同一目录下，所以不用写具体路径
    rowField = []

    with open(filename, 'r') as file:
        for x in file:
            pgnRow = []
            pgnRowArray = []
            if sPGN in x:
                x = x.strip('\n')
                pgnRow.append(x.split('\t'))
                pgnRowArray = np.array(pgnRow)[0]
                spn = pgnRowArray[1]
                Bpostion = pgnRowArray[2]
                unit = pgnRowArray[3]
                length = (pgnRowArray[3].split())[0]
                resolution = pgnRowArray[4].split()[0]
                offset = pgnRowArray[5].split()[0]
                if 'ASCII' in resolution:
                    if spn == '234':
                        hexStr1 = (''.join(hexArray[1:]))
                    elif spn == '1635':
                        hexStr1 = (''.join(hexArray[4:19]))
                    elif sPGN == '65259':
                        hexStr1 = (''.join(hexArray[:]))
                    # print('hex', hexStr1)
                    spnResult = bytes.fromhex(hexStr1).decode().strip(b'\x00'.decode())
                    if spn == '586':
                        spnResult = spnResult.split('*')[0]
                    elif spn == '587':
                        spnResult = spnResult.split('*')[1]
                    elif spn == '588':
                        spnResult = spnResult.split('*')[2]
                    print('asccii spn值:', spnResult)
                else:
                    if 'byte' in unit:
                        postionArr = []
                        hexArray2 = []
                        if '-' in Bpostion:
                            postionArr = np.array(Bpostion.split('-'))
                            hexArray2 = hexArray[(int(postionArr[0])-1):int(postionArr[1])]
                            # print(hexArray2)
                            hexArray2 = hexArray2[::-1]
                            # print(hexArray2)
                            # string1 = '0x' + (''.join(hexArray2))
                            hexStr1 = (''.join(hexArray2))
                            decNum = int(hexStr1, 16)
                            if '/' in resolution:
                                spnResult = decNum * float(Fraction(resolution)) + float(offset)
                            else:
                                spnResult = decNum * float(resolution) + float(offset)

                            print('spn值:', spnResult)

                        else:
                            hexStr1 = hexArray[int(Bpostion)-1]
                            decNum = int(hexStr1, 16)
                            spnResult = decNum * float(resolution) + float(offset)
                            print('spn值:', spnResult)

                    else:
                        # print('Bpostion===',Bpostion)
                        Bytepos, binpos = Bpostion.split('.')
                        hexStr1 = hexArray[int(Bytepos)-1]
                        decNum = int(hexStr1, 16)
                        binNum = '{:08b}'.format(decNum) 
                        # print('binNum:',binNum, type(binNum))                         
                        binbegin = -int(binpos)+1
                        binend = -int(length) + (-int(binpos)) + 1
                        # print(binbegin)
                        # print(binend)

                        if int(binpos) == 1:
                            binResult = binNum[binend:]
                            # print(binNum[binend:])
                        else:
                            binResult = binNum[binend:binbegin]
                            # print(binNum[binend:binbegin])

                        if 'states' in pgnRowArray[4]:
                            binResult = int(binResult, 2)
                            print('spn值:', binResult)
                        else:
                            binResult = int(binResult, 2) * float(resolution) + float(offset)
                            print('spn值:', binResult)


# -----------------------------------------快照参数的对比----------------------------------------------------------------


# 1、读取一条HB，把快照参数保存为字典dict{spn:value}
# json它是一种基于文本，独立于语言的轻量级数据交换格式,一般接口传输数据用。
def read_hb_json(fn_json):
    with open(fn_json, encoding='UTF-8') as fd:
        # load：针对文件句柄，将json格式的字符转换为dict，从文件中读取 (将string转换为dict)
        data = json.load(fd)
        # dumps：将dict转换为string (易传输)
        df = pd.read_json(json.dumps(data['Snapshots'][0]['Parameter']))
        # print(df.head())
        # print(len(df))
        # hb_cd_dict是HB里面的快照参数dict
        hb_cd_dict = {c['Name'].lower(): c['Value'] for _, c in df.iterrows()}
        # print(hb_cd_dict)
        # 获取到HB的Occurrence_Date_Time根据此时间和tbox的采集逻辑来确定查找报文的时间范围
        utc_time_str = data['Occurrence_Date_Time']
        utc_format = '%Y-%m-%dT%H:%M:%S.%fZ'
        utc_dt = datetime.datetime.strptime(utc_time_str, utc_format)
        global start_time, end_time
        # occurrence_date_time是UTC时间+8小时
        occurrence_date_time = utc_dt + timedelta(hours=8)
        print(f'\n{fn_json} occurrence_date_time == {occurrence_date_time}')
        # 此条HB报文开始时间和结束时间，可以多找几分钟
        start_time = (occurrence_date_time + timedelta(seconds=-25)).strftime("%H:%M:%S.%f")
        end_time = (occurrence_date_time + timedelta(seconds=25)).strftime("%H:%M:%S.%f")

    # 读取同步性sheet
    df = pd.read_excel('CDTestCase.xlsx', sheetname='同步性', header = 0)

    # print(df.head())
    # print(df.columns)

    global spn_dict, pgn_dict, pgn_list
    spn_dict = {}
    pgn_dict = {}
    pgn_list = []

    for index, row in df.iterrows():
        pgn = row[3]
        if np.isnan(pgn):
            pass
        else:
            pgn = int(row[3])
            if pgn not in pgn_list:
                pgn_list.append(pgn)

    # pgn_list.pop(0)
    # print(pgn_list)

    # test_dict = {c['SDK Interface Format v 2.603.000']: c['J1939\nSPN'] for _, c in df.iterrows()} # noqa
    test_dict = {c[1].lower(): c[4] for _, c in df.iterrows()} # noqa
    # 根据SDK同步性sheet，test_dict = {快照参数名称：spn}
    # print(test_dict) 
    # 此循环是把HB的快照参数和SDK里面的快照参数进行遍历，找到对应的参数，组合为最新的字典spn_dict = {spn:value}
    for k in test_dict:
        if k in hb_cd_dict.keys():
            spn_dict[int(test_dict[k])] = hb_cd_dict[k]

    # print(spn_dict)


# 2、开始处理报文
# DataFrame是一种表格型数据结构，它含有一组有序的列，每列可以是不同的值
def read_hex_csv():

    # 读取报文hex_message.csv
    df = pd.read_csv('hex_message.csv')
    df['NewTime'] = df['Time']
    df['NewTime'] = df['NewTime'].apply(lambda x: x.rsplit('.', 1)[0])

    global start_time, end_time, pgn_list
    print(start_time)
    print(end_time)
    df = df[(df['Time'] > str(start_time)) & (df['Time'] < str(end_time))]
    spn_pgn_dict = {}
    for pgn in pgn_list:
        # print(pgn)
        df_pgn = df[(df['PGN'] == pgn)]
        if len(df_pgn) == 0:
            continue
            
        # print(df_pgn.head())
        not_found = True
        for index, row in df_pgn.iterrows():
            spn_pgn_dict = hex_pgn_to_spn(str(pgn), row['Data'])
            # if pgn == 54016:
            #     print(f'===54016 {spn_pgn_dict}====')

            equal = True
            for spn_key in spn_pgn_dict.keys():
                if spn_key not in spn_dict.keys():
                    equal = False
                    break
                # 这4个spn的值是ASCII的直接对比
                if spn_key in (1635, 234, 586, 587) :
                    if spn_pgn_dict[spn_key] != spn_dict[spn_key]:
                        equal = False
                        break
                else:
                    # HB上报的数据小数点不统一，所以需要把HB的值和报文算出来的值转换为float，2个值的差的绝对值小于0.01是
                    # if float(spn_pgn_dict[spn_key]) != float(spn_dict[spn_key]):
                    if (abs(float(spn_pgn_dict[spn_key]) - float(spn_dict[spn_key])) >= 0.01):
                        equal = False
                        break
            if equal:
                not_found = False

        fail_spn_dict = {}
        for item in spn_pgn_dict:
            if item in spn_dict:
                fail_spn_dict[item] = spn_dict[item]
        if not_found:
            print(f'\tcompare pgn {pgn} fail hex={spn_pgn_dict} <==> HB={fail_spn_dict}')
        else:
            print(f'\tcompare pgn {pgn} success hex={spn_pgn_dict} <==> HB={fail_spn_dict}')




# 根据pgn和报文，计算出对应spn的值，select_spn_dict = {spn:value}
def hex_pgn_to_spn(pgn, hexStr):
    sPGN = str(pgn)
    # print(len(hexStr))
    # 把16进制的报文，转换为字节数组
    hexArray = np.array(hexStr.split())
    # print(len(hexArray))
    # txt文件和当前脚本在同一目录下，所以不用写具体路径
    filename = 'rules.txt' 
    rowField = []
    select_spn_dict = {}
    # 读取spn计算规则的文本文件，一行一行读取
    with open(filename, 'r') as file:
        for x in file:
            pgnRow = []
            pgnRowArray = []
            # 先判断传过来的pgn是否是此行的pgn
            if sPGN in x:
                # 去掉结尾的换行符
                x = x.strip('\n')
                # 按照制表符'\t'切割字符串，得到的结果构成一个数组
                pgnRow.append(x.split('\t'))
                pgnRowArray = np.array(pgnRow)[0]
                # 找出此数组里面的对应的spn，字节位置，长度，resolution，offset
                spn = pgnRowArray[1]
                Bpostion = pgnRowArray[2]
                unit = pgnRowArray[3]
                length = (pgnRowArray[3].split())[0]
                resolution = pgnRowArray[4].split()[0]
                offset = pgnRowArray[5].split()[0]
                # 1、需要转换为ASCII的参数，单独处理
                if 'ASCII' in resolution:
                    if spn == '234':
                        hexStr1 = (''.join(hexArray[1:]))
                    elif spn == '1635':
                        hexStr1 = (''.join(hexArray[4:19]))
                    elif sPGN == '65259':
                        hexStr1 = (''.join(hexArray[:]))
                    # print('...', type(hexStr1))
                    # print(spn)
                    # print('hex', hexStr1)
                    # if spn == '1635':
                    #     spnResult = bytes.fromhex(hexStr1.strip('0')+'0').decode()
                    # else:
                    #     spnResult = bytes.fromhex(hexStr1.strip('0')).decode()
                    spnResult = bytes.fromhex(hexStr1.strip('0')).decode()


                    if spn == '586':
                        spnResult = spnResult.split('*')[0]
                    elif spn == '587':
                        spnResult = spnResult.split('*')[1]
                    select_spn_dict[int(spn)] = str(spnResult)
                    # print('select_spn_dict:', select_spn_dict)
                else:
                    # 2、以byte为单位的
                    if 'byte' in unit:
                        postionArr = []
                        hexArray2 = []
                        if '-' in Bpostion:
                            postionArr = np.array(Bpostion.split('-'))
                            hexArray2 = hexArray[(int(postionArr[0])-1):int(postionArr[1])]
                            # print(hexArray2)
                            hexArray2 = hexArray2[::-1]
                            # print(hexArray2)
                            # string1 = '0x' + (''.join(hexArray2))
                            hexStr1 = (''.join(hexArray2))
                            decNum = int(hexStr1, 16)
                            if '/' in resolution:
                                spnResult = decNum * float(Fraction(resolution)) + float(offset)
                            else:
                                spnResult = decNum * float(resolution) + float(offset)

                            select_spn_dict[int(spn)] = str(spnResult)
                            # print('select_spn_dict:', select_spn_dict)
                        else:
                            hexStr1 = hexArray[int(Bpostion)-1]
                            decNum = int(hexStr1, 16)
                            spnResult = decNum * float(resolution) + float(offset)
                            select_spn_dict[int(spn)] = str(spnResult)
                            # print('select_spn_dict:', select_spn_dict)
                    else:
                        # 3、以为bit为单位的
                        # print('Bpostion===',Bpostion)
                        Bytepos, binpos = Bpostion.split('.')
                        hexStr1 = hexArray[int(Bytepos)-1]
                        decNum = int(hexStr1, 16)
                        binNum = '{:08b}'.format(decNum)
                        # print('binNum:',binNum, type(binNum))
                        binbegin = -int(binpos)+1
                        binend = -int(length) + (-int(binpos)) + 1
                        # print(binbegin)
                        # print(binend)

                        if int(binpos) == 1:
                            binResult = binNum[binend:]
                            # print(binNum[binend:])
                        else:
                            binResult = binNum[binend:binbegin]
                            # print(binNum[binend:binbegin])

                        if 'states' in pgnRowArray[4]:
                            binResult = int(binResult, 2)
                            select_spn_dict[int(spn)] = str(binResult)
                            # print('select_spn_dict:', select_spn_dict)
                        else:
                            binResult = int(binResult, 2) * float(resolution) + float(offset)
                            select_spn_dict[int(spn)] = str(binResult)
                            # print('select_spn_dict:', select_spn_dict)
    # print('*********', select_spn_dict)
    return select_spn_dict
                    

def compare_hb_hex(type_cd):
    json_string = type_cd + '_CD.json'
    HBFC_to_xlsx(type_cd)
    read_hb_json(json_string)
    read_hex_csv()

if __name__ == '__main__':

    # read_message_trc()
    # get_hb_json()
    # calTrim()
    compare_hb_hex('HB')
# 
    # hex_pgn_to_spn(54016, 'DF EF 93 6F 53 43 39 34 33 31 35 00 00 00 00 00 00 00 00 00')  

    # PGNHexToDec(54016, 'DF EF 93 6F 53 43 39 34 33 31 35 00 00 00 00 00 00 00 00 00')  
    # hex_pgn_to_spn()  
    # hex_to_hb('HB')
